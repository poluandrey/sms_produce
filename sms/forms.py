from django.core.exceptions import ValidationError
from django.forms import ModelForm, FileField, ModelMultipleChoiceField
from django.contrib.admin.widgets import AdminDateWidget
from openpyxl.reader.excel import load_workbook

from sms.models import Broadcast, Text, Sender


class AddBroadcastForm(ModelForm):
    prefix_file = FileField(label='Upload Prefix File', help_text='Upload a CSV file containing prefixes.', required=False)
    text = ModelMultipleChoiceField(queryset=Text.objects.all(), required=True)
    sender = ModelMultipleChoiceField(queryset=Sender.objects.all(), required=True)

    class Meta:
        model = Broadcast
        fields = [
            'name',
            'comment',
            'is_active',
            'start_date',
            'end_date',
            'phone_number_length',
            'channel_login',
            'channel_password',
            'total_sms_count'
        ]
        widgets = {
            'start_date': AdminDateWidget(),
            'end_date': AdminDateWidget(),
        }

    def clean(self):
        cleaned_data = super(AddBroadcastForm, self).clean()
        return cleaned_data

    def clean_prefix_file(self):
        prefix_file = self.cleaned_data.get('prefix_file')
        if prefix_file:
            try:
                workbook = load_workbook(filename=prefix_file)
                sheet = workbook.active
                prefixes = [row[0] for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True)]
                self.cleaned_data['prefixes'] = prefixes
            except Exception as e:
                raise ValidationError(f"Error processing file: {e}")
        return prefix_file
